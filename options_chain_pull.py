from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs
import requests
import ssl
import sys
import tdameritrade.auth #added as40183
import urllib
import urllib3 #as40183

from sys import argv
import pymysql.cursors
import datetime
import dateutil.relativedelta
import calendar
import time
import json
import ast
import pandas
import sqlite3

import string
import xlwt
import openpyxl

KEY = 'STOCKTIPS'
# Arguements
in_file = r'C:\Anupam\market\stock_options_api-master\trading_api\tdameritrade\my_programs\data\program_in.txt'
out_file=r'C:\Anupam\market\stock_options_api-master\trading_api\tdameritrade\my_programs\data\program_out.txt'
script='C:/Anupam/market/stock_options_api-master/trading_api/tdameritrade/my_programs/options_chain_pull.py'

debug = 'true'
f_in = open(in_file)
equity_list = f_in.readlines()

equity_list = [l.replace('\n','') for l in equity_list]

f_out = open(out_file,'w')
print ('EQUITY | CMP | 52WkRange', file=f_out)

#sqlite3 connection
connection = sqlite3.connect('C:\Anupam\Technical\sqlite\db\mydb.db')
cursor = connection.cursor()

create_sql = """CREATE TABLE IF NOT EXISTS chain (
                equity text NOT NULL,
                symbol text NOT NULL,
                cmp real NOT NULL, --added from stocks
                _52WkRange text NOT NULL, --added from stocks
                strikePrice real NOT NULL,
                last real NOT NULL,
                bid real NOT NULL,
                ask real NOT NULL,
                bidSize real NOT NULL,
                askSize real NOT NULL,
                totalVolume real NOT NULL,
                volatility real NOT NULL,
                putCall text NOT NULL,
                inTheMoney text NOT NULL,
                daysToExpiration int NOT NULL,
                timeValue real NOT NULL,
                theoreticalVolatility real NOT NULL
                );"""

drop_sql = "DROP TABLE CHAIN"
select_sql = "SELECT * FROM CHAIN"
delete_sql = "DELETE FROM CHAIN"

if (debug == 'true'):
    print ('create_sql==',create_sql)
    print ('delete_sql==',delete_sql)

#cursor.execute(drop_sql)
cursor.execute(create_sql)
cursor.execute(delete_sql)
connection.commit()

cursor.execute(select_sql)
row=cursor.fetchall()
print (row)

# Declare
#start = datetime.now()
args_list = []
count = str(250)
myFormat = "%Y-%m-%d"
today = datetime.date.today()
rd = dateutil.relativedelta.relativedelta(days=1, weekday=dateutil.relativedelta.FR)
next_friday = today + rd
if (debug == 'true'):
    print ('today=',today)
    print('next_friday=',str(next_friday))

#debug: Remove comment to use expiration of a future date
next_friday=today+datetime.timedelta(days=17)
print('next_friday=', str(next_friday))

#debug starts
#equity='AAPL'
count=40
start_date=next_friday

#active_day variables start - syncup from excel_pull

#CUSTOMIZATION BLOCK starts
debug='false'
skip_days=0 #set to 0 if placing order today; update to 1 if need for tomorrow+day-after-tomorrow
#CUSTOMIZATION BLOCK ends

curr_date = datetime.date.today() + datetime.timedelta(days=skip_days)

if (curr_date.isoweekday() == 6):
    curr_date = curr_date + datetime.timedelta(days=2)
elif (curr_date.isoweekday() == 7):
    curr_date = curr_date + datetime.timedelta(days=1)


if curr_date.isoweekday() in set((5, 6)):
    next_date = curr_date + datetime.timedelta(days=8 - curr_date.isoweekday())
else:
    next_date = curr_date + datetime.timedelta(days=1)

print (curr_date, calendar.day_name[curr_date.weekday()], curr_date.isoweekday())
print (next_date, calendar.day_name[next_date.weekday()], next_date.isoweekday())

active_day_today = calendar.day_name[curr_date.weekday()]
active_day_tomorrow = calendar.day_name[next_date.weekday()]
print (active_day_today, active_day_tomorrow)
#active_day variables end

# for NEXT WEEK FRIDAY DEBUG only
#next_friday = next_friday + datetime.timedelta(days=7)
#start_date = start_date + datetime.timedelta(days=7)
#print("NEXT WEEK next_friday-start_date", next_friday, start_date)
# END OF NEXT WEEK DEBUG

for equity in equity_list:
#EQUITY STOCK CODE
    time.sleep(1.01)
    equity, mkt_time = equity.split(",")
    equity = equity.strip()
    print('equity=', equity)
    start_equity = datetime.datetime.now()
    url = 'https://api.tdameritrade.com/v1/marketdata/'+equity+'/quotes?apikey='+KEY
    #url1 = 'https://api.tdameritrade.com/v1/marketdata/AAPL/quotes?apikey=STOCKTIPS'
    r = requests.get(url)
    payload = r.json()

    if (debug=='true'):
        print(url)
        print ('r=',r)
        print ('r.text=',r.text)
        print ('payload=',payload)

    equity = payload[equity]['symbol']
    cmp = payload[equity]['regularMarketLastPrice'] #lastPrice
    _52WkLow = round(payload[equity]['52WkLow'])
    _52WkHigh = round(payload[equity]['52WkHigh'])
    if (debug=='true'):
        print ('equity=',equity)
        print ('cmp=',cmp)
        print ('EQUITY | CMP | 52WkRange', file=f_out)
        print (equity, '|', cmp, '|', _52WkLow, '-', _52WkHigh, file=f_out)


#OPTION CHAIN CODE
    url = 'https://api.tdameritrade.com/v1/marketdata/chains?apikey=' + KEY + \
          '&symbol=' + equity + '&contractType=' + 'PUT' + '&range=OTM' + '&fromDate=' + \
          str(start_date) + '&toDate=' + str(next_friday) + '&strikeCount=' + str(count) # + '&strike<170.0'
    r = requests.get(url)  # <Response [200]>
    payload = r.json()
    if (debug == 'true'):
        print('URL==', url)
        print(r.text)
        print(payload)

    symbol = payload['symbol']

    # Get Puts
    for keyy, valuee in payload["putExpDateMap"].items():
        d = datetime.datetime.strptime(keyy, "%Y-%m-%d:%f")
        ex_date = d.strftime(myFormat)
        for key, value in valuee.items():
            for v in value:
                args = [ v['symbol'], payload["symbol"], v['strikePrice'], v['last'], v['bid'], v['ask'], v['bidSize'], v['askSize'], v['totalVolume'], v['volatility'], v['putCall'], ex_date, v['inTheMoney'], v['daysToExpiration'], v['timeValue'], v['theoreticalVolatility'] ]
                if (debug == 'true'):
                    print (v['strikePrice'] ,'CMP=', float(cmp))
                if (v['strikePrice'] < float(cmp)):
                    args_list.append(args)
                    if (debug == 'true'):
                        print ('args_list=',args_list)

                    insert_sql = "INSERT INTO CHAIN (" \
                    + " equity, symbol, cmp, _52WkRange, strikePrice, last, bid, ask, bidSize, askSize, totalVolume, volatility, putCall, inTheMoney, daysToExpiration, timeValue, theoreticalVolatility " \
                    + ") values ('" \
                    + payload['symbol'] + "','" \
                    + v['symbol'] + "'," \
                    + str(cmp) + "," \
                    + str("'" + str(_52WkLow) + "-" + str(_52WkHigh)) + "'" + "," \
                    + str(v['strikePrice']) + "," \
                    + str(v['last']) + "," \
                    + str(v['bid']) + "," \
                    + str(v['ask']) + "," \
                    + str(v['bidSize']) + "," \
                    + str(v['askSize']) + ","  \
                    + str(v['totalVolume']) + "," \
                    + str(v['volatility']) + ",'" \
                    + str(v['putCall']) + "','" \
                    + str(v['inTheMoney']) + "'," \
                    + str(v['daysToExpiration']) + "," \
                    + str(v['timeValue']) + "," \
                    + str(v['theoreticalVolatility']) \
                    + ")"

                    if (debug == 'true'):
                        print ('insert_sql==',insert_sql)

                    cursor.execute(insert_sql)
                    connection.commit()

# FINAL RESULT SQLs
wbkName_out = r'C:\Anupam\market\consolidated_excel_data.xlsx'
wbk_out = openpyxl.load_workbook(wbkName_out)
wks_out = wbk_out[active_day_today+'-'+active_day_tomorrow]

#WRITE OUTPUT TO EXCEL
select_sql1 = "select distinct equity, market_time from chain order by equity;"
print ('select_sql1=',select_sql1)

print ("-----------------", file=f_out)

select_sql2 = "select distinct equity, cmp, _52WkRange from chain order by equity;"
print ('select_sql2=',select_sql2)
cursor.execute(select_sql2)
rows = cursor.fetchall()
idx=2
#wks_out.cell(row=1, column=3).value = " ".join(["EQUITY" , " | " , "CMP" , "|" , "52WkRange"])
wks_out.cell(row=1, column=2).value = "52WkRange"
wks_out.cell(row=1, column=5).value = "CMP"
for row in rows:
    if (debug == 'true'):
        print('select_sql2:', row[0], "|" ,row[1], "|", row[2])
    #wks_out.cell(row=idx, column=3).value = " ".join([str(row[0]) , "|" , str(row[1]) , "|" , str(row[2])])
    wks_out.cell(row=idx, column=2).value = str(row[2])
    wks_out.cell(row=idx, column=5).value = row[1]
    idx+= 1
    if (debug == 'true'):
        print ('select_sql2:idx=',idx)

print ("-----------------", file=f_out)

select_sql3 = "select equity, strikeprice, bid, round(bid*100/strikeprice,2) prem_per from chain " +  \
              "where equity||strikeprice in (select equity||max(strikeprice) from chain group by equity) order by equity;"
print ('select_sql3=',select_sql3)
cursor.execute(select_sql3)
rows = cursor.fetchall()
idx=2
#wks_out.cell(row=1, column=4).value = " ".join(["EQUITY" , " | " , "STRIKEPRICE" , "|" , "BID", "|", "PREM_PCT"])
wks_out.cell(row=1, column=3).value = "EQUITY"
wks_out.cell(row=1, column=6).value = "STRIKEPRICE"
wks_out.cell(row=1, column=7).value = "BID"
wks_out.cell(row=1, column=8).value = "PREM_PCT"


for row in rows:
    print(row[0], "|" ,row[1], "|", row[2], "|", row[3])
    #wks_out.cell(row=idx, column=4).value = " ".join([str(row[0]) , "|" , str(row[1]) , "|" , str(row[2]), "|" , str(row[3])])
    wks_out.cell(row=idx, column=3).value = str(row[0])
    wks_out.cell(row=idx, column=6).value = row[1]
    wks_out.cell(row=idx, column=7).value = row[2]
    wks_out.cell(row=idx, column=8).value = row[3]
    idx+= 1
    if (debug == 'true'):
        print ('select_sql3:idx=',idx)

print ("-----------------", file=f_out)

select_sql4 = "select equity, strikeprice, round(((cmp-strikeprice)*-100/cmp),1) prc_diff, bid, round(bid*100/strikeprice,1) prem_per from chain a " + \
              "where bid>=0.05 and (prc_diff <=-5 and prc_diff >= -12) or (prc_diff <= -14 and prc_diff >= -20) " + \
              "order by equity, prc_diff;"
print ('select_sql4=',select_sql4)
cursor.execute(select_sql4)
rows = cursor.fetchall()
idx=2
wks_out.cell(row=1, column=9).value = " ".join(["EQUITY" , " | " , "STRIKEPRICE" , "|", "PCT_DIFF", "|" , "BID", "|", "PREM_PCT"])
for row in rows:
    new_eq = row[0]
    if (debug == 'true'):
        print(row[0], "|" ,row[1], "|", row[2], "|", row[3], "|", row[4], file=f_out)
    wks_out.cell(row=idx, column=9).value = " ".join([str(row[0]) , "|" , str(row[1]) , "|" , str(row[2]), "|" , str(row[3]), "|" , str(row[4])])
    idx += 1
    if (debug == 'true'):
        print('idx=', idx)
    prev_eq = row[0]
    if (prev_eq != new_eq):
        print ("---", file=f_out)


wbk_out.save(wbkName_out)
wbk_out.close
