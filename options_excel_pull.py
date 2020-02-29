import pandas as pd
import xlrd
import sqlite3
from pandas.io import sql
import datetime
import calendar

import openpyxl

#sqlite3 connection
connection = sqlite3.connect('C:\Anupam\Technical\sqlite\db\mydb.db')
cursor = connection.cursor()

out_file = r'C:\Anupam\market\stock_options_api-master\trading_api\tdameritrade\my_programs\data\program_in.txt'
f_out = open(out_file,'w')


'''
create_sql = """CREATE TABLE IF NOT EXISTS cboe_data(
                "index" INTEGER,
                equity text NULL,
                equity_desc text NULL,
                product_type text NULL
                );"""


create_sql = """CREATE TABLE IF NOT EXISTS Tuesday_TDA(
                "index" INTEGER,
                equity text NULL,
                equity_desc text NULL,
                market_time text NULL,
                confirmed text NULL
                );"""

#cursor.execute(create_sql)
connection.commit()

#END OF SQL CONNECTION
'''

#CUSTOMIZATION BLOCK starts
debug='true'
#sheets_list = ['Custom_TDA','Custom_Nasdaq']
sheets_list = ['CBOE_data', 'Monday_TDA', 'Monday_Nasdaq', 'Tuesday_TDA', 'Tuesday_Nasdaq', 'Wednesday_TDA', 'Wednesday_Nasdaq', 'Thursday_TDA', 'Thursday_Nasdaq', 'Friday_TDA', 'Friday_Nasdaq',
               'Monday_Chameleon', 'Tuesday_Chameleon', 'Wednesday_Chameleon', 'Thursday_Chameleon', 'Friday_Chameleon']

skip_days=0 #set to 0 if placing order today; update to 1 if need for tomorrow+day-after-tomorrow
#CUSTOMIZATION BLOCK ends

curr_date = datetime.date.today() + datetime.timedelta(days=skip_days)

if (curr_date.isoweekday() == 6):
    if (debug=='true'):
        print ('Today_Saturday:', curr_date.isoweekday())
    curr_date = curr_date + datetime.timedelta(days=2)
elif (curr_date.isoweekday() == 7):
    if (debug == 'true'):
        print ('Today_Sunday:', curr_date.isoweekday())
    curr_date = curr_date + datetime.timedelta(days=1)


if curr_date.isoweekday() in set((5, 6)):
    if (debug == 'true'):
        print ('Today either Sat or Sun')
    next_date = curr_date + datetime.timedelta(days=8 - curr_date.isoweekday())
else:
    if (debug == 'true'):
        print ('Today is weekday:', curr_date.isoweekday())
    next_date = curr_date + datetime.timedelta(days=1)

print (curr_date, calendar.day_name[curr_date.weekday()], curr_date.isoweekday())
print (next_date, calendar.day_name[next_date.weekday()], next_date.isoweekday())

active_day_today = calendar.day_name[curr_date.weekday()]
active_day_tomorrow = calendar.day_name[next_date.weekday()]

if (debug == 'true'):
    print (active_day_today, active_day_tomorrow)

wbkName_in = "C:\Anupam\market\consolidated_excel_data.xlsx"
wbk_in = pd.ExcelFile(wbkName_in) #reading file; CBOE_data
if (debug == 'true'):
    print ('Complete sheets list:',wbk_in.sheet_names)
#sheets_list = xl.sheet_names

if (debug == 'true'):
    print('Currently processing sheets list:', sheets_list)

for v_sheet_name in sheets_list:
    #if ( 'CBOE_data' in v_sheet_name or active_day_today in v_sheet_name or active_day_tomorrow in v_sheet_name or 'Custom' in v_sheet_name):
    todays_sheets_list = [ 'CBOE_data', active_day_today + '_Chameleon', active_day_today + '_Nasdaq', active_day_today + '_TDA', active_day_tomorrow + '_Nasdaq', active_day_tomorrow + '_TDA']
    if ( v_sheet_name in todays_sheets_list ):
        print ('Active: v_sheet_name=',v_sheet_name.upper())
        if (v_sheet_name.upper() == 'CBOE_DATA'):
            df = wbk_in.parse(v_sheet_name, usecols="A,C:D")
            print ("------85 -------")
        elif ('TDA' in v_sheet_name):
            df = wbk_in.parse(v_sheet_name, usecols="A,B,G")
            print("------88 -------")
        elif ('Nasdaq' in v_sheet_name):
            df = wbk_in.parse(v_sheet_name, usecols="J,K,L")
            print("------91 -------")
        elif ('Zacks' in v_sheet_name):
            df = wbk_in.parse(v_sheet_name, usecols="A,B,D")
            print("------94 -------")
        elif (v_sheet_name == active_day_today + '_Chameleon'):
            df = wbk_in.parse(v_sheet_name, usecols="A,B,D")
        if (debug == 'true'):
            print("------DF-------", df)

        if ( len(df) >0 ):
            if (v_sheet_name.upper() == 'CBOE_DATA'):
                df.columns = ['equity', 'equity_desc', 'product_type']
            elif ('TDA' in v_sheet_name or 'Zacks' in v_sheet_name or 'Nasdaq' in v_sheet_name or 'Chameleon' in v_sheet_name):
                df.columns = ['equity', 'equity_desc', 'market_time']

        if (debug == 'true'):
            print('Shape', df.shape)
            print('-------------------------')
            print('Number of rows', len(df))
            print('-------------------------')
            print('Column headers', df.columns)
            print('-------------------------')
            print('Data types', df.dtypes)
            print('-------------------------')
            print('Index', df.index)
            print('-------------------------')

        #print(df)
        #print (len(df.axes[0]))

        #Load worksheet to table
        cursor.execute ("delete from " + v_sheet_name + ";") #not really needed since we could use replace in to_sql below.
        print("Old record count to be deleted: " + str(cursor.rowcount))

        df.to_sql(v_sheet_name, connection, if_exists='replace') #changed form replace

        if (len(df) > 0):
            delete_sql = "DELETE FROM "+ v_sheet_name + " where equity is null or equity='';"
            if (debug == 'true'):
                print('delete_sql==', delete_sql)
            cursor.execute(delete_sql)

        #cursor.execute(drop_sql)

        if (('Zacks' in v_sheet_name or 'Chameleon' in v_sheet_name) and len(df) >0 ):
            update_sqla = "UPDATE " + v_sheet_name + " set market_time='After' where lower(market_time)='amc';"
            update_sqlb = "UPDATE " + v_sheet_name + " set market_time='Before' where lower(market_time)='bmo';"
            if (debug == 'true'):
                print (update_sqla, "\n",update_sqlb)

            cursor.execute(update_sqla)
            cursor.execute(update_sqlb)

        elif ( v_sheet_name.upper() == 'CBOE_DATA'):
            update_sql = "UPDATE " + v_sheet_name + " set equity=replace( equity,'/', '');"
            if (debug == 'true'):
                print("update_sql= ", update_sql)
            cursor.execute(update_sql)
#        else:
#            print ("THIS IS NOT IN SCOPE TODAY: ", v_sheet_name)
    connection.commit()

#file = "C:\Anupam\market\consolidated_excel_data.xlsx"
#wbk_in.save(wbkName_in)
wbk_in.close

#WRITE EQUITY-DESC-MARKET_TIME TO EXCEL
wbkName_out = r'C:\Anupam\market\stock_options_api-master\trading_api\tdameritrade\my_programs\data\trial.xlsx'
wbk_out = openpyxl.load_workbook(wbkName_in)
wks_out = wbk_out[active_day_today+'-'+active_day_tomorrow]

#Clear sheet contents first
r=1000
c=12
for row in range(1,r):
    for col in range(1,c):
        wks_out.cell(row=row, column=col).value=None


equity_list_sql = "select distinct stk.equity, replace( trim(cboe.equity_desc,'*'), ',' , '') equity_desc, stk.market_time from ( " + \
                " select equity, market_time||'-" + active_day_today + "' market_time from " + active_day_today + "_tda where market_time ='After' " + \
                " union " + \
                " select equity, market_time||'-" + active_day_today + "' market_time from " + active_day_today + "_nasdaq where market_time ='After' " + \
                " union " + \
                " select equity, market_time||'-" + active_day_tomorrow + "' market_time from " + active_day_tomorrow + "_tda where market_time ='Before' " + \
                " union " + \
                " select equity, market_time||'-" + active_day_tomorrow + "' market_time from " + active_day_tomorrow + "_nasdaq where market_time ='Before' " + \
                " union " + \
                " select equity, market_time||'-" + active_day_today + "' market_time from " + active_day_today + "_Chameleon where market_time ='After' " + \
                " union " + \
                " select equity, market_time||'-" + active_day_tomorrow + "' market_time from " + active_day_today + "_Chameleon where market_time ='Before' " + \
                " ) stk, cboe_data cboe " + \
                " where stk.equity = cboe.equity" + \
                " and cboe.product_type = 'Equity'" + \
                " order by stk.equity; "



print ("equity_list_sql==",equity_list_sql)
cursor.execute(equity_list_sql)
rows = cursor.fetchall()
for row in rows:
    print(row[0], "," ,row[1], file=f_out)


idx=2
wks_out.cell(row=1, column=1).value = "Equity Name"
wks_out.cell(row=1, column=4).value = "Market_Time"
for row in rows:
    print(row[0], "|" ,row[1])
    wks_out.cell(row=idx, column=1).value = str(row[1])
    wks_out.cell(row=idx, column=4).value = str(row[2])
    idx+= 1
    if (debug == 'true'):
        print ('idx=',idx)

wbk_out.save(wbkName_in)
wbk_out.close

